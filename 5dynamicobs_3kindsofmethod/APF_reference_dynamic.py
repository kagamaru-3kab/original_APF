"""
conventional method(reference)
経路をグラフに表示
2D,3Dポテンシャルも表示
静的障害物に対して一時的なゴールを作成
dynamicbs を作成
"""
import numpy as np
from matplotlib import pyplot as plt
from matplotlib import patches as pat
from matplotlib import cm
from openpyxl import Workbook
import copy

class goal():                          
    def __init__(self, locate, area):
        """   
        #Content:set up goal 
        diameter_size: size 実際の大きさ
        locate: goal position[x,y]  
        attracted_area: effective area [int a]影響を与える範囲
        """
        self.diameter_size = 0.1        
        self.locate_goal = locate
        self.attracted_area = area
    
    def get_locate(self, locate):
        self.locate_goal = locate

class obstacles():
    def __init__(self, obstacles = None):  
        """   
        Content: set up obstacles
        obstacles: ID and position array([x1,y1][x2,y2]..)
        """
        self.locate_obstacles = obstacles
    
    def update_locate_obs(self):
        """   
        Content: update location dynamic obstacles ,assign dynamic obs id and displacement
        obstacles: ID and position array([x1,y1][x2,y2]..)
        """
        dynamic_obs_id = [0,1]
        self.obs_velocity_vector = [[-0.1,-0.1],[-0.1, -0.1]]#needed proposal dimension

        for d in range(len(dynamic_obs_id)): 
           self.locate_obstacles[dynamic_obs_id[d]][0] += self.obs_velocity_vector[d][0]
           self.locate_obstacles[dynamic_obs_id[d]][1] += self.obs_velocity_vector[d][1]
        
class vehicles():
    def __init__(self, init_locate): 
        """   
        Content: set up obstacles
        diameter_size: size 実際の大きさ
        speed: vehicles
        """
        self.diameter = 0.5
        self.speed    = 0.1
        self.init_locate = init_locate
        self.locate_vehicles = init_locate
    
    def get_locate(self, locate):     #update vehicle position
        self.locate_vehicles = locate
        return self.locate_vehicles
    
class calc_APF():    
    """
    content: calculate artifitial potential field
    dist_v2goal: include distance between vehicle and goal
    attracte_k, repulse_k: include  coefficient goal and obstacle
    vehicle_speed: update distance  
    repulse_area: effective area of obstacle
    """               
    def __init__(self,vehicles_speed):
        self.dist_v2goal = None     
        self.attracte_k  = 10
        self.repulse_k   = 0.3
        self.vehicles_speed = vehicles_speed
        self.repulsed_area = 5
    
    def calc_goal_dist_theta(self, locate_goal, locate_vehicles): #calculate distance between goal and vehicle
        self.dist_v2goal = np.sqrt((locate_goal[0]-locate_vehicles[0])**2+(locate_goal[1]-locate_vehicles[1])**2)
        #self.theta_goal= np.arctan2(locate_goal[1]-locate_vehicles[1], locate_goal[0]-locate_vehicles[0])
        #self.goalanddistandtheta = [self.dist_v2goal, self.theta_goal]

    def calc_obs_dist_theta(self, locate_obs, locate_vehicles): #calculate distance between obstacles and vehicle
        self.dist_v2obs = []
        self.theta_obs  = []
        self.obs_distandtheta = []
        for id in range(len(locate_obs)):
            self.dist_v2obs.append(np.sqrt((locate_obs[id][0]-locate_vehicles[0])**2+(locate_obs[id][1]-locate_vehicles[1])**2))
            #self.theta_obs.append(np.arctan2(locate_obs[id][1]-locate_vehicles[1], locate_obs[id][0]-locate_vehicles[0]))                  

    def calc_attractive_force(self, locate_goal, locate_vehicles): #calculate attractive_force 
        self.calc_goal_dist_theta(locate_goal, locate_vehicles)
        if self.dist_v2goal == 0:
            self.attforce = -1
        else:
            self.attforce = -1*self.attracte_k/self.dist_v2goal
    
    def calc_repulsive_force(self, locate_obs, locate_vehicles): #calculate repulse_force
        if len(locate_obs) == None:
            self.repforce = [0]
        else:
            self.repforce = [0]*len(locate_obs)
            self.calc_obs_dist_theta(locate_obs, locate_vehicles)
            for id in range(len(self.dist_v2obs)):
                if self.dist_v2obs[id] <= 0:
                    self.repforce[id] = 1
                elif self.repulsed_area >= self.dist_v2obs[id]:
                    self.repforce[id] = 1*self.repulse_k/self.dist_v2obs[id]
                else: self.repforce[id] = 0 

    def calc_sum_potentialforce(self, locate_goal, locate_vehicles, locate_obs): #calculate sum attractive_force
        self.calc_attractive_force(locate_goal, locate_vehicles)    
        self.calc_repulsive_force(locate_obs, locate_vehicles)
        self.total_pot = self.attforce
        for id in range(len(self.repforce)):
            self.total_pot = self.total_pot + self.repforce[id]
        return self.total_pot
    
    def detect_tairyuu(self, partialdiffer_x, partialdiffer_y):# detect that vehicles stucking
        if np.sign(partialdiffer_x) == 1:
            flagx = 1
            print("flagx =1")
        elif np.sign(partialdiffer_x) == -1:
            flagx = 0
            print("flagx =0")
        if np.sign(partialdiffer_y) == 1:
            flagy = 1
        elif np.sign(partialdiffer_y) == -1:
            flagy = 0
        tairyux = 0
        tairyuy = 0
        print("before,after",self.before_flagx,flagx)
        if self.before_flagx != flagx:
            self.countx += 1 
            print("countx ===", self.countx)
            if self.countx >= 20:
                print("x ga tairyuu")
                tairyux = 1
        if self.before_flagy != flagy:
            self.county += 1
            if self.county >= 20:
                print("y ga tairyuu")
                tairyuy = 1
        if tairyux == 1 and tairyuy == 1:
            print("tairyuuuuuu")
        self.before_flagx = flagx
        self.before_flagy = flagy

    def route_creater(self, locate_goal, locate_vehicles, locate_obs): #route_creater calculates movement amount from partitial differential of potential 
        current_potential = self.calc_sum_potentialforce(locate_goal, locate_vehicles, locate_obs)
        delt_poten_x = self.calc_sum_potentialforce(locate_goal, [locate_vehicles[0]+self.vehicles_speed, locate_vehicles[1]], locate_obs)
        delt_poten_y = self.calc_sum_potentialforce(locate_goal, [locate_vehicles[0] ,locate_vehicles[1]+self.vehicles_speed], locate_obs)
        partialdiffer_x = delt_poten_x - current_potential
        partialdiffer_y = delt_poten_y - current_potential
        synthesis_v = np.sqrt(partialdiffer_x**2+partialdiffer_y**2)
        partialdiffer_x /=synthesis_v/self.vehicles_speed*-1   #正規化らしい
        partialdiffer_y /=synthesis_v/self.vehicles_speed*-1
        
        return partialdiffer_x, partialdiffer_y

class plot_path(): #plot vehicle trajectory
    """
    set param  ploting path
    search about matplotlib with Internet
    """
    def __init__(self, start_position , goal_position):
        self.fig = plt.figure(figsize=(7,7))
        self.ploting_path = self.fig.add_subplot(111)
        self.ploting_path.set_xlabel('X-distance: m')
        self.ploting_path.set_ylabel('Y-distance: m')
        self.goal_position = goal_position
        self.start_position = start_position
        self._initialize_fig()

    def _initialize_fig(self):
        self.graph_range = [0, 150]
        plt.xlim([self.graph_range[0], self.graph_range[1]])
        plt.ylim([self.graph_range[0], self.graph_range[1]])
        self.ploting_path.plot(self.start_position[0], self.start_position[1], "*r")
        self.ploting_path.plot(self.goal_position[0], self.goal_position[1], "*b")
        circle_goal = pat.Circle(xy=(self.goal_position[0], self.goal_position[1]), radius=2, fc ="y")
        self.ploting_path.add_patch(circle_goal)
        self.plot_obs(obs.locate_obstacles)
    
    def plot_obs(self, obstacles): #plot set obstacles area
        r_area = APF.repulsed_area
        for i in range(len(obstacles)):
            circle_obs = pat.Circle(xy=(obstacles[i][0],obstacles[i][1]), radius= r_area, fc ="b", alpha = 0.3)
            self.ploting_path.add_patch(circle_obs)
            self.ploting_path.plot(obstacles[i][0],obstacles[i][1], "xk")
    
    def plot_vehicle(self, locate_vehile): #plot locate_vehicle 
        self.ploting_path.plot(locate_vehile[0], locate_vehile[1], ".r")

    def calc_all_potential(self): #calculate potential whole area
        pot = []
        for y_for_pot in range(self.graph_range[0], self.graph_range[1] + 1):
            tmp_pot = []
            for x_for_pot in range(self.graph_range[0], self.graph_range[1] + 1):
                veh1.locate_vehicles = [x_for_pot, y_for_pot]
                potential = APF.calc_sum_potentialforce(fgoal.locate_goal, veh1.locate_vehicles, obs.locate_obstacles)
                tmp_pot.append(potential)
            pot.append(tmp_pot)
        pot = np.array(pot)
        return pot
    
    def plot_3d_potential(self): #plot potential force to 3D gragh
        fig3d = plt.figure(figsize=(6,6), facecolor= "w")
        ax3d = fig3d.add_subplot(111, projection="3d")
        ax3d.tick_params(labelsize=7)    # 軸のフォントサイズ
        ax3d.set_xlabel("x", fontsize=10)
        ax3d.set_ylabel("y", fontsize=10)
        ax3d.set_zlabel("poten", fontsize=10)
        pot = self.calc_all_potential()
        x, y = np.meshgrid(np.arange(self.graph_range[0], self.graph_range[1] + 1), np.arange(self.graph_range[0], self.graph_range[1] + 1))
        surf = ax3d.plot_surface(x, y, pot, rstride=1, cstride=1, cmap= cm.coolwarm)
        plt.show()

class temporary_goal():
    """
    content: find the nearest obstacle on line from vehicle to goal and set temporary goal beside target obstacle
    """
    def __init__(self):
        self.tmp_goal_locate = None
        self.nearest_obs = None

    def calc_line_from_start2goal(self, locate_goal, locate_vehicles): #calc coefficient of line which pass start and goal 
        self.a = locate_goal[1]-locate_vehicles[1]
        self.b = locate_vehicles[0]-locate_goal[0]
        self.c = locate_vehicles[1]*locate_goal[0]-locate_goal[1]*locate_vehicles[0]

    def draw_line_from_start2goal(self, locate_goal, locate_vehicles): #plot_line which pass start and goal  
        print("goal,veh1 position in draw line",locate_goal, locate_vehicles)
        startp = int(locate_vehicles[0])
        endp = int(locate_goal[0])
        if self.b != 0 and self.a != 0 :
            self.typeline = 0
            for x in range(startp, endp):
                y = (-1*self.a*x-self.c)/self.b
                path_fig.ploting_path.plot(x, y, ".k")
        elif self.b == 0 and self.a != 0:
            self.typeline = 1
            for y in range(startp, endp):
                x = locate_vehicles[0] 
                path_fig.ploting_path.plot(x, y, ".k")

        elif self.a == 0 and self.b != 0:
            self.typeline = 2
            for x in range(startp, endp):
                y = locate_vehicles[1] 
                path_fig.ploting_path.plot(x, y, ".k")
        else:
            print("set goal and start on same point")

    def detect_obs_ontheline(self, locate_obs): #judge which obstacles on the line
        if len(locate_obs) != 0:
            dist_from_line = [0]*len(locate_obs)
            self.register_id_obs = []
            self.Denominator = np.sqrt(self.a**2+self.b**2)
            print("obstacles exist")
            #print("a, b, c, self.Denominator", self.a, self.b, self.c, self.Denominator)
            for i in range(len(locate_obs)):
                if (fgoal.locate_goal[0]-veh1.locate_vehicles[0]) >= 0 :
                    if locate_obs[i][0] >= veh1.locate_vehicles[0] and locate_obs[i][0] <= fgoal.locate_goal[0]:
                        Numerator = abs(self.a*locate_obs[i][0]+self.b*locate_obs[i][1]+self.c)
                    else:
                        Numerator = -1
                elif (fgoal.locate_goal[0]-veh1.locate_vehicles[0]) < 0 :
                    if locate_obs[i][0] <= veh1.locate_vehicles[0] and locate_obs[i][0] >= fgoal.locate_goal[0]:
                        Numerator = abs(self.a*locate_obs[i][0]+self.b*locate_obs[i][1]+self.c)
                    else:
                        Numerator = -1
                else:
                    Numerator = -1
                print("i, numera",i, Numerator)
                if Numerator > 0:
                    dist_from_line[i] = Numerator/self.Denominator
                elif Numerator == 0:
                    dist_from_line[i] = 0
                else:
                    print("Numerator is -1 because target_goal is out of range")
                    dist_from_line[i] = APF.repulsed_area + 1
                print("id + dist_from_line", i, dist_from_line[i])
                if dist_from_line[i] < APF.repulsed_area:
                    self.register_id_obs.append(i)
                    print("found obs on line")
            if len(self.register_id_obs) == 0:
                self.register_id_obs = None 
                print("obs exist but not on line")      
        elif len(locate_obs) == 0:
            self.register_id_obs = None
            print(" obstacles no exist")
        
        if self.register_id_obs == None:
            print("registerID is None")
        else:
            print("ゴールまでの線上のobsは",len(self.register_id_obs),"個だ")
    
    def find_nearest_obs(self): #find nearest obs on the line from vehicle
        if self.register_id_obs  != None:
            for i, id_obs in enumerate(self.register_id_obs):
                if i == 0:
                    self.nearest_obs = [id_obs, APF.dist_v2obs[id_obs]]
                else:
                    if self.nearest_obs[1] > APF.dist_v2obs[id_obs]:
                        self.nearest_obs = [id_obs, APF.dist_v2obs[id_obs]]
            print("車両最近傍のobsは[id, dist_fromvehicle]", self.nearest_obs)
        else:
            self.nearest_obs = None

    def calc_tempgoal_locate(self, Dfortemp, temp_locate): #ref https://qiita.com/tydesign/items/36b42465b3f5086bd0c5
        delta_fromobs = 1
        aD = self.slope_nomal*Dfortemp
        a2b2 = self.slope_nomal**2 + 1**2
        rr = APF.repulsed_area**2
        contain_root = np.sqrt((a2b2*rr) - Dfortemp**2)
        x1 = (aD - contain_root)/a2b2
        x1 += temp_locate[0] -delta_fromobs
        x2 = (aD + contain_root)/a2b2
        x2 += temp_locate[0] +delta_fromobs
        y1 = self.slope_nomal*x1 +self.intercept_nomal
        y2 = self.slope_nomal*x2 +self.intercept_nomal
        p1 = [x1,y1]
        p2 = [x2,y2]
        Numerator_p1 = abs(self.a*p1[0]+self.b*p1[1]+self.c)
        Numerator_p2 = abs(self.a*p2[0]+self.b*p2[1]+self.c)
        near_goal_dist = Numerator_p1/self.Denominator
        if near_goal_dist > Numerator_p2/self.Denominator:
            return p2
        else: 
            return p1 
        
        
    def set_temporary_goal(self): #set temporary goal
        if self.nearest_obs !=  None:
            temporary_goal_locate = [obs.locate_obstacles[self.nearest_obs[0]][0],obs.locate_obstacles[self.nearest_obs[0]][1]]
            #print("temporary_goal_before_changed",temporary_goal_locate)
            self.slope_nomal = self.b/self.a                                                        #法線の傾きを求めてる
            self.intercept_nomal = temporary_goal_locate[1]-(self.slope_nomal*temporary_goal_locate[0])#法線の切片を求めてる
            Dfortemp = abs(self.slope_nomal*temporary_goal_locate[0]-temporary_goal_locate[1]+self.intercept_nomal)
            #print("slope nomal and intercept",self.slope_nomal, self.intercept_nomal)
            near_tmpgoal = self.calc_tempgoal_locate(Dfortemp, temporary_goal_locate)
            temporary_goal_locate = near_tmpgoal          
            path_fig.ploting_path.plot([0], temporary_goal_locate[1], "xy")
            return temporary_goal_locate
        else:
            print("not need to change goal")
            temporary_goal_locate = None
            return temporary_goal_locate 
        
    def draw_nomalline(self): #plot_line which pass start and goal  
        if self.nearest_obs != None:
            for x in range(50):
                y = self.slope_nomal*x + self.intercept_nomal
                #print("nomal x, y: ", x, y)
                if y >= 0 or y <= 50:
                    path_fig.ploting_path.plot(x, y,".g")
                else:
                    break

    def integrate_temporarygoal(self,locate_goal,locate_vehicles, locate_obs): #integrate this class function
        self.calc_line_from_start2goal(locate_goal,locate_vehicles)
        self.draw_line_from_start2goal( locate_goal, locate_vehicles)
        self.detect_obs_ontheline(locate_obs)
        self.find_nearest_obs()
        self.temp_goal = self.set_temporary_goal()
        self.draw_nomalline()
        #print("changed_temp_goal",self.temp_goal)


class guide_point():
    """
    Content: implement reference method on paper named Energy Efficient Local Path Planning Algorithm
    Based on Predictive Artificial Potential Field

    self.length_precede : max precede distance
    self.numberofpoints : number of precede points
    self.interval_points : interval each points
    self.guide_point : contain points location
    self.guide_point[0,0] self.guide_point[0,1] : initial location 
    """
    def __init__(self):
        self.length_precede = 16
        self.numberofpoints = 4
        self.interval_points = self.length_precede/self.numberofpoints
        self.guide_point = np.zeros((self.numberofpoints,2))
        #print("guide_point_shape",self.guide_point)
        self.guide_point[0,0] = veh1.locate_vehicles[0]
        self.guide_point[0,1] = veh1.locate_vehicles[1]
        temp_goal.calc_line_from_start2goal(fgoal.locate_goal,veh1.locate_vehicles)
        temp_goal.draw_line_from_start2goal( fgoal.locate_goal, veh1.locate_vehicles)
        temp_goal.detect_obs_ontheline(obs.locate_obstacles)
        self.count = 0
        self.temp_goal_point = fgoal.locate_goal
    

    def set_guide_point(self):#calc location of precede points
        i = 0
        count_calc_guidepoint = 0
        flagtostopprecede = False
        self.Denominator = np.sqrt(temp_goal.a**2+temp_goal.b**2)
        update_interval = self.interval_points
        print("do set guide point")
        while i < self.numberofpoints  :
                x, y = APF.route_creater(fgoal.locate_goal, self.guide_point[i], obs.locate_obstacles)
                #print(x,y,self.guide_point)
                self.guide_point[i,0] = x+self.guide_point[i,0]
                self.guide_point[i,1] = y+self.guide_point[i,1]
                #print("self.guide_point",self.guide_point[i-1,0],self.guide_point[i-1,1])
                self.dist_guide2vehicle = np.sqrt((self.guide_point[i,0]-veh1.locate_vehicles[0])**2+(self.guide_point[i,1]-veh1.locate_vehicles[1])**2)
                self.dist_guide2fgoal = np.sqrt((fgoal.locate_goal[0]-self.guide_point[i,0])**2+(fgoal.locate_goal[1]-self.guide_point[i,1])**2)
                self.dist_v2finalgoal = np.sqrt((fgoal.locate_goal[0]-veh1.locate_vehicles[0])**2+(fgoal.locate_goal[1]-veh1.locate_vehicles[1])**2)
                #print("countcalc_guideponint",count_calc_guidepoint)
                count_calc_guidepoint += 1
                if self.dist_guide2fgoal < self.interval_points:
                    flagtostopprecede = True
                    break
                if count_calc_guidepoint > 1000:
                    print("guidepoints are stuck")
                    self.guide_point[i,0] = self.guide_point[i,0] + -0.1
                    count_calc_guidepoint = 0

                if self.dist_guide2vehicle > update_interval and self.dist_v2finalgoal > self.dist_guide2fgoal :
                    #print("distv2fgoal - distguide2fgoal",self.dist_v2finalgoal - self.dist_guide2fgoal)
                    print("now calc guide point number is",i,"dist,x,y",self.dist_guide2vehicle,self.guide_point[i,0], self.guide_point[i,1])
                    update_interval = update_interval + self.interval_points
                    path_fig.ploting_path.plot(self.guide_point[i,0], self.guide_point[i,1], "*y")
                    i += 1
                    count_calc_guidepoint = 0
        print("guidepoint locete\n", self.guide_point)

        for i in range(self.numberofpoints): 
            #print("range num",i)           
            
            Numerator = abs(temp_goal.a*self.guide_point[i,0]+temp_goal.b*self.guide_point[i,1]+temp_goal.c)              
            self.dist_guide2line = Numerator/self.Denominator
            if i == 0:
                max_dist = self.dist_guide2line
                max_number = i
            elif i > 0:
                if max_dist <= self.dist_guide2line:
                    max_dist = self.dist_guide2line
                    max_number = i
        if flagtostopprecede:
            self.temp_goal_point = None
            path_fig.ploting_path.plot(fgoal.locate_goal[0], fgoal.locate_goal[1], "xr")
        else:       
            self.temp_goal_point = self.guide_point[max_number]
            print("tempgoal,maxdist,max_num",self.temp_goal_point,max_dist,max_number)
            path_fig.ploting_path.plot(self.temp_goal_point[0], self.temp_goal_point[1], "xr")

        if Numerator == -1:
            self.temp_goal_point = fgoal.locate_goal
            print("numerator is -1 so set finalgoal")
        #print("self.temp_goal_point  comforming",self.temp_goal_point)
        return self.temp_goal_point


def main(): 
    if APF.dist_v2goal == None:   #clac distance to reach goal at first time
        APF.calc_goal_dist_theta(fgoal.locate_goal, veh1.locate_vehicles)
        APF.calc_obs_dist_theta(obs.locate_obstacles, veh1.locate_vehicles)
        #print("\nif goal dist is None calc once:", APF.dist_v2goal)
        #print("show all dist from vehicle to every obs:", APF.dist_v2obs)
        #print("veh1 position now",veh1.locate_vehicles)
        #temp_goal.integrate_temporarygoal(fgoal.locate_goal, veh1.locate_vehicles, obs.locate_obstacles)
        temp_goal.temp_goal = guide.set_guide_point()
        #print("temp_goal.temp_goal in main func is ",temp_goal.temp_goal)
        if np.all(temp_goal.temp_goal != None):
            target_goal = temp_goal.temp_goal
            judgereach_finalgoal = True
        elif temp_goal.temp_goal == None:
            target_goal = fgoal.locate_goal
            judgereach_finalgoal = False
        print("target_goal",judgereach_finalgoal)
    if APF.dist_v2goal != None:
        while (APF.dist_v2goal - veh1.diameter) > fgoal.diameter_size:  #iterate until reach goal    
            partialdiffer_x, partialdiffer_y = APF.route_creater(target_goal, veh1.locate_vehicles, obs.locate_obstacles)
            coefficient = 3
            partialdiffer_x, partialdiffer_y =coefficient*partialdiffer_x,coefficient*partialdiffer_y 
            veh1.locate_vehicles = [veh1.locate_vehicles[0]+partialdiffer_x, veh1.locate_vehicles[1]+partialdiffer_y]
            obs.update_locate_obs()
            excel_data_route.append(veh1.locate_vehicles)
            obs_locatecopy = copy.deepcopy(obs.locate_obstacles)
            obs_trajectry.append(obs_locatecopy)
            if all(target_goal) == None:
                target_goal_record.append([0,0])
            else:
                target_goal_record.append(target_goal)
            if any(dist < 2 for dist in APF.dist_v2obs):
                hitcount[0] +=1
                print("APF.dist_v2obs の中に 1 以上の要素が存在します。",hitcount[0])
            path_fig._initialize_fig()
            path_fig.plot_vehicle(veh1.locate_vehicles)
            plt.pause(0.04)
            plt.cla()
        if judgereach_finalgoal:
            APF.dist_v2goal = None
            plt.cla()
            print("reach tempo goal but not reach final goal")
            #print("received repforce",APF.repforce)
            return main()
        else:
            print("finish and show 3d")
            #path_fig.plot_3d_potential()

        wb = Workbook()
        ws = wb.active
        previous_max_column = 0
        previous_max = 0
        label_list = ['V_traject_X', 'V_traject_Y', 'obs_locate_x', 'obs_locate_y','obs_speed_x','obs_speed_y','target_goal_x','target_goal_y','hitcount']
        for col_index, value in enumerate(label_list, start=1):
                    cell = ws.cell(row=1, column=col_index, value=value)

        for row_index, row in enumerate(excel_data_route, start=2):
            for col_index, value in enumerate(row, start=1):
                cell = ws.cell(row=row_index, column=col_index, value=value)
                if previous_max < col_index + previous_max_column:
                    previous_max = col_index + previous_max_column
        previous_max_column = previous_max
        # obs_trajectry の処理
        previous_max = 0
        for row_index, row_data in enumerate(obs_trajectry, start=2):
            for col_index, value in enumerate(row_data[0], start=1):
                cell = ws.cell(row=row_index, column=previous_max_column+col_index, value=value)
                if previous_max < col_index + previous_max_column:
                    previous_max = col_index + previous_max_column
        previous_max_column = previous_max

        previous_max = 0
        for row_index, row_data in enumerate(obs_trajectry, start=2):
            for col_index, value in enumerate(row_data[1], start=1):
                cell = ws.cell(row=row_index, column=previous_max_column+col_index, value=value)
                if previous_max < col_index + previous_max_column:
                    previous_max = col_index + previous_max_column
        previous_max_column = previous_max
        # obs.obs_velocity_vector の処理
        previous_max = 0
        for row_index, row in enumerate(obs.obs_velocity_vector, start=2):
            for col_index, value in enumerate(row, start=1):
                cell = ws.cell(row=row_index, column= previous_max_column+col_index, value=value)
                if previous_max < col_index + previous_max_column:
                    previous_max = col_index + previous_max_column
        previous_max_column = previous_max

        previous_max = 0
        for row_index, row in enumerate(target_goal_record, start=2):
            for col_index, value in enumerate(row, start=1):
                cell = ws.cell(row=row_index, column=previous_max_column+col_index, value=value)
                if previous_max < col_index + previous_max_column:
                    previous_max = col_index + previous_max_column
        previous_max_column = previous_max
        
        cell = ws.cell(row=2, column= previous_max_column+1, value=hitcount[0])
                
        # Excelファイルに保存
        wb.save('4research_APF_fig\excel_date/refer_mix_2d_ff.xlsx')


if __name__ == '__main__':
    fgoal = goal([70,70],10)
    obs = obstacles([[70,75],[60,60]])
    veh1 = vehicles([0,0])
    APF = calc_APF(veh1.speed)
    path_fig = plot_path(veh1.locate_vehicles, fgoal.locate_goal)
    temp_goal = temporary_goal()
    guide = guide_point()
    excel_data_route = []
    excel_data_route.append(veh1.init_locate)
    obs_trajectry = []
    hitcount = [0]
    target_goal_record = []
    main()